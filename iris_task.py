import openpyxl
from fastapi.responses import FileResponse
from fastapi import FastAPI
from pydantic import BaseModel
import uvicorn
from fastapi import Query


wb = openpyxl.load_workbook("C:/Users/Sharmeen/Downloads/capbudg.xlsx")
app = FastAPI()

sheet = wb.active
# print(sheet['A1'].value)
#
# for row in sheet.iter_rows(min_row=2,values_only=True):
#     print(row)

print(wb.sheetnames)
for name in wb.sheetnames:
    print(wb[name].tables.keys())

import re

# def parse_numeric(cell_value):
#     if isinstance(cell_value, (int, float)):
#         return cell_value
#     if isinstance(cell_value, str):
#         cleaned = cell_value.replace('$', '').replace(',', '').strip()
#         try:
#             if cleaned.endswith('%'):
#                 return float(cleaned.strip('%')) / 100
#             return float(cleaned)
#         except ValueError:
#             return None
#     return None

def parse_numeric(value):
    if isinstance(value, (int, float)):
        return value
    if isinstance(value, str):
        cleaned = re.sub(r'[^\d.\-]', '', value)
        try:
            return float(cleaned)
        except ValueError:
            return None
    return None


@app.get("/list_tables/")
async def list_tables():
    table_names = list(sheet.tables.keys())
    return {"tables":table_names}

# class TableDetail(BaseModel):
#     table_name: str

@app.get("/get_table_details/")
async def get_table_details(table_name: str = Query(..., description="Name of the Excel table")):
    if table_name not in sheet.tables:
        return {"error": "Table not found"}

    table_obj = sheet.tables[table_name]
    table_range = table_obj.ref
    min_col, min_row, max_col, max_row = openpyxl.utils.range_boundaries(table_range)
    row_names = [
        sheet.cell(row=row, column=min_col).value
        for row in range(min_row + 1, max_row + 1)
    ]

    return {"row_names": row_names}
    # row_names = [cell.value for cell in sheet[1]]
    # for name in row_names:
    #     return {"row_names":row_names}

# class RowSum(BaseModel):
#     table_name: str
#     row_name: str

@app.get("/row_sum/")
async def row_sum(table_name: str = Query(...,description="Enter name of the Excel table"), row_name: str = Query(...,description="Enter name of the excel row"),):
    if table_name not in sheet.tables:
        return {"error": "Table not found"}

    table_range = sheet.tables[table_name].ref
    min_col, min_row, max_col, max_row = openpyxl.utils.range_boundaries(table_range)
    print("Table name received:", table_name)
    print("Row name received:", row_name)
    print("Table ref range:", table_range)
    print("Row scan from:", min_row + 1, "to", max_row)

    for row in range(min_row + 1, max_row + 1):
        if sheet.cell(row=row, column=min_col).value == row_name:
            values = [
                parse_numeric(sheet.cell(row=row, column=col).value)
                for col in range(min_col + 1, max_col + 1)
            ]
            values = [v for v in values if v is not None]
            return {"parsed_values": values, "row_sum": sum(values)}

    return {"error": "Row name not found"}

if __name__ == "__main__":
    uvicorn.run(app, host="0.0.0.0", port=9090,log_level='info')






